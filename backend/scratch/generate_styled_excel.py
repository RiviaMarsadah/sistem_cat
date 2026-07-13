import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Create workbook
wb = openpyxl.Workbook()

# Setup Sheet 1: Detail Pengujian
ws1 = wb.active
ws1.title = "Detail Pengujian"
ws1.views.sheetView[0].showGridLines = True

# Common styles
font_family = "Segoe UI"
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

# Colors
navy_header_fill = PatternFill(start_color='0F1F3D', end_color='0F1F3D', fill_type='solid')
module_header_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
pass_fill = PatternFill(start_color='D4EDDA', end_color='D4EDDA', fill_type='solid')
zebra_fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

# Set Title block for Sheet 1
ws1.merge_cells("A1:E1")
ws1["A1"] = "LAPORAN DETAIL PENGUJIAN BLACK BOX - SISTEM ATEKA (RIVIA CAT)"
ws1["A1"].font = Font(name=font_family, size=14, bold=True, color='0F1F3D')
ws1["A1"].alignment = Alignment(horizontal='left', vertical='center')
ws1.row_dimensions[1].height = 30

# Column Headers (Exactly: No, Skenario Pengujian, Hasil yang Diharapkan, Hasil yang Didapat, Hasil)
headers = [
    "No", 
    "Skenario Pengujian", 
    "Hasil yang Diharapkan", 
    "Hasil yang Didapat", 
    "Hasil"
]

header_row_idx = 3
for col_idx, header in enumerate(headers, 1):
    cell = ws1.cell(row=header_row_idx, column=col_idx, value=header)
    cell.font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
    cell.fill = navy_header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws1.row_dimensions[header_row_idx].height = 28

# Modules and Scenarios Data (Without TS-XX-XX codes)
modules_data = [
    {
        "name": "Modul 1: Autentikasi & Otorisasi",
        "scenarios": [
            ("Login dengan email & password terdaftar sesuai role (Guru)", 
             "Sistem memvalidasi kredensial dan mengarahkan user ke Dashboard Guru.", 
             "Berhasil login dan halaman dialihkan ke Dashboard Guru.", "Sesuai"),
            ("Login dengan password salah", 
             "Sistem menampilkan pesan kesalahan 'Password salah'.", 
             "Muncul pesan error 'Password salah' di layar.", "Sesuai"),
            ("Login menggunakan Google OAuth tertaut", 
             "Sistem mengotentikasi akun dan mengarahkan ke dashboard yang sesuai tanpa meminta password.", 
             "Otentikasi Google berhasil dan langsung masuk ke dashboard.", "Sesuai"),
            ("Otorisasi bypass URL halaman admin", 
             "Sistem menolak akses, menampilkan pesan otorisasi tidak valid, dan mengarahkan kembali (redirect).", 
             "Akses ditolak oleh middleware auth dan diarahkan kembali ke dashboard siswa.", "Sesuai")
        ]
    },
    {
        "name": "Modul 2: Sinkronisasi Data (SIJUWAN Sync)",
        "scenarios": [
            ("Analisis perbedaan data sebelum sinkronisasi", 
             "Sistem membandingkan data lokal dan eksternal, lalu menampilkan rincian data baru, perubahan, dan konflik.", 
             "Tampil rincian perbedaan data beserta indikator perubahan secara dinamis.", "Sesuai"),
            ("Eksekusi sinkronisasi dengan data besar", 
             "Proses berjalan lancar dalam mode batch, menampilkan progress bar real-time via Server-Sent Events (SSE) tanpa timeout.", 
             "Sinkronisasi selesai 100% dan progress bar ter-update real-time via SSE.", "Sesuai"),
            ("Deteksi konflik duplikasi email saat sinkronisasi", 
             "Sistem mendeteksi konflik, melewatkan (skip) data tersebut dari pembuatan user baru, dan mencatat log konflik.", 
             "Data siswa dilewatkan dari sync and log konflik mencantumkan alasan duplikasi email.", "Sesuai")
        ]
    },
    {
        "name": "Modul 3: Manajemen Akademik (Jurusan, Kelas, Angkatan, Mapel)",
        "scenarios": [
            ("Tambah Jurusan/Prodi dengan kode prodi valid", 
             "Jurusan baru berhasil disimpan di database.", 
             "Jurusan tersimpan dan tampil pada tabel daftar jurusan.", "Sesuai"),
            ("Validasi duplikasi Kode Jurusan/Prodi", 
             "Sistem menolak penyimpanan dan memunculkan pesan validasi 'Kode Prodi sudah digunakan'.", 
             "Proses simpan diblokir dan muncul pesan error duplikasi kode.", "Sesuai"),
            ("Tambah Kelas baru terikat tingkat & jurusan", 
             "Sistem memetakan kelas ke Jurusan RPL dengan nama kelas 'XII RPL 2' secara otomatis.", 
             "Kelas berhasil dibuat dan terikat pada jurusan RPL.", "Sesuai"),
            ("Tambah Mata Pelajaran baru", 
             "Mata pelajaran tersimpan di database dan siap digunakan untuk paket soal.", 
             "Mapel sukses tersimpan dan tampil pada daftar mapel.", "Sesuai"),
            ("Tambah Angkatan baru", 
             "Angkatan baru tersimpan dan dapat ditautkan ke siswa.", 
             "Angkatan tersimpan dengan tahun 2026.", "Sesuai")
        ]
    },
    {
        "name": "Modul 4: Manajemen Soal (Bank Soal & Koleksi)",
        "scenarios": [
            ("Tambah Bank Soal Koleksi", 
             "Koleksi bank soal berhasil dibuat sebagai wadah pengelompokan butir soal.", 
             "Koleksi berhasil dibuat dan tampil di halaman bank soal guru.", "Sesuai"),
            ("Pembuatan soal Pilihan Ganda (Pilgan)", 
             "Soal tersimpan dengan opsi lengkap dan kunci jawaban terikat secara benar.", 
             "Soal pilgan tersimpan dengan sukses.", "Sesuai"),
            ("Pembuatan soal Pilihan Ganda Kompleks (PG Kompleks)", 
             "Sistem menyimpan soal dan menormalisasikan urutan kunci jawaban menjadi string terpisah koma.", 
             "Soal kompleks berhasil disimpan dengan kunci jawaban 'A,C,D'.", "Sesuai"),
            ("Pembuatan soal Pilihan Ganda Kategori (PG Kategori)", 
             "Sistem menyimpan soal beserta kolom pernyataan dan merekam status kunci masing-masing baris secara terurut.", 
             "Soal kategori tersimpan beserta data kunci kebenaran per baris.", "Sesuai")
        ]
    },
    {
        "name": "Modul 5: Penjadwalan Ujian (Jadwal Wizard)",
        "scenarios": [
            ("Membuat Jadwal Ujian Resmi", 
             "Jadwal ujian berhasil disimpan beserta relasi kelas penerima dan token masuk di-generate otomatis.", 
             "Jadwal sukses dibuat dengan Token Masuk dan Token Checkout ter-generate.", "Sesuai"),
            ("Validasi rentang waktu mulai/selesai yang tidak valid", 
             "Wizard menolak proses simpan dan menampilkan validasi error waktu.", 
             "Form menampilkan pesan error 'Waktu selesai harus setelah waktu mulai'.", "Sesuai"),
            ("Auto-generate Token Ujian dan Token Checkout", 
             "Token terisi otomatis di form jadwal dan tersimpan dengan aman.", 
             "Kedua token terbuat otomatis berupa 6 digit kapital acak.", "Sesuai"),
            ("Menghubungkan Paket Soal Guru ke Jadwal Resmi Admin (Labeling Resmi)", 
             "Jadwal ujian terbarui dengan label 'Admin - [Nama Guru] Resmi' dan paket ujian ter-link.", 
             "Status jadwal berubah menjadi terhubung dan label resmi tersemat di dashboard monitoring.", "Sesuai")
        ]
    },
    {
        "name": "Modul 6: Pelaksanaan Ujian Siswa (Mobile App)",
        "scenarios": [
            ("Masuk Ujian menggunakan Token valid", 
             "Aplikasi memverifikasi token dan mengarahkan siswa ke Attempt Screen dengan timer berjalan.", 
             "Berhasil masuk ke ruang ujian dan timer pengerjaan dimulai.", "Sesuai"),
            ("Masuk Ujian diluar jam/jadwal atau kelas salah", 
             "Sistem menolak akses dan menampilkan pesan dialog error yang informatif.", 
             "Muncul pesan error bahwa ujian belum dimulai atau kelas siswa tidak terdaftar.", "Sesuai"),
            ("Auto-save progress jawaban siswa berkala secara async", 
             "Aplikasi secara async mengirim status jawaban ke database server tanpa reload halaman/aplikasi.", 
             "Status jawaban terubah di server menjadi 'dijawab' dan progress aman.", "Sesuai"),
            ("Pemicu Kiosk Mode untuk mengunci layar siswa", 
             "Aplikasi mengaktifkan Kiosk Mode (mengunci layar, memblokir tombol home/recent apps).", 
             "Kiosk Mode aktif dan siswa terkunci di dalam aplikasi ujian.", "Sesuai"),
            ("Selesai ujian lebih awal dengan Token Checkout", 
             "Sistem memproses submit, menghitung nilai akhir secara otomatis, dan menutup sesi ujian siswa.", 
             "Ujian disubmit secara sah dan nilai siswa berhasil dihitung di server.", "Sesuai"),
            ("Force-submit otomatis saat waktu habis (timer 00:00)", 
             "Aplikasi secara otomatis mengirimkan jawaban akhir siswa ke server tanpa memerlukan Token Checkout.", 
             "Jawaban terkirim otomatis, sesi ditutup, dan halaman dialihkan ke hasil ujian.", "Sesuai")
        ]
    },
    {
        "name": "Modul 7: Analisis & Rekap Hasil Ujian (Guru/Admin)",
        "scenarios": [
            ("Rekap Nilai Ujian Siswa per Kelas", 
             "Sistem menampilkan tabel nilai akhir seluruh siswa di kelas tersebut beserta rata-rata, nilai tertinggi, dan terendah.", 
             "Tabel rekap nilai tampil lengkap dengan visualisasi ringkasan kartu statistik di bagian atas.", "Sesuai"),
            ("Analisis Butir Soal", 
             "Sistem menampilkan jumlah siswa yang menjawab benar, salah, atau kosong untuk masing-masing opsi pilihan ganda.", 
             "Sebaran jawaban tampil detail dalam bentuk diagram dan tabel analisis.", "Sesuai"),
            ("Ekspor Rekap Hasil Ujian ke format Excel (xlsx)", 
             "Sistem meng-generate file spreadsheet (.xlsx) berisi daftar nilai siswa secara instan.", 
             "File Excel ter-download dengan format yang rapi dan data nilai lengkap.", "Sesuai")
        ]
    }
]

current_row = 4
global_counter = 1

for mod in modules_data:
    # Write Module subheader row
    ws1.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=5)
    mod_cell = ws1.cell(row=current_row, column=1, value=mod["name"])
    mod_cell.font = Font(name=font_family, size=11, bold=True, color='0F1F3D')
    mod_cell.fill = module_header_fill
    mod_cell.alignment = Alignment(horizontal='left', vertical='center')
    
    # Set border for merged module row
    for col in range(1, 6):
        c = ws1.cell(row=current_row, column=col)
        c.border = thin_border
    ws1.row_dimensions[current_row].height = 24
    
    current_row += 1
    
    # Write scenarios
    for idx, (title, expected, actual, status) in enumerate(mod["scenarios"], 1):
        row_vals = [global_counter, title, expected, actual, status]
        
        # Zebra striping for readability
        row_fill = zebra_fill if idx % 2 == 0 else PatternFill(fill_type=None)
        
        for col_idx, val in enumerate(row_vals, 1):
            cell = ws1.cell(row=current_row, column=col_idx, value=val)
            cell.font = Font(name=font_family, size=10)
            cell.border = thin_border
            if idx % 2 == 0:
                cell.fill = row_fill
                
            # Alignment rules
            if col_idx in [1, 5]: # No, Hasil
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
            # Style status cell
            if col_idx == 5:
                cell.font = Font(name=font_family, size=10, bold=True, color='155724')
                cell.fill = pass_fill
                
        ws1.row_dimensions[current_row].height = 28
        global_counter += 1
        current_row += 1

# Setup column widths for Sheet 1
column_widths_s1 = {
    'A': 6,   # No
    'B': 45,  # Skenario Pengujian
    'C': 45,  # Hasil yang Diharapkan
    'D': 45,  # Hasil yang Didapat
    'E': 12   # Hasil
}
for col_letter, width in column_widths_s1.items():
    ws1.column_dimensions[col_letter].width = width


# Setup Sheet 2: Rekapitulasi
ws2 = wb.create_sheet(title="Rekapitulasi")
ws2.views.sheetView[0].showGridLines = True

# Title block
ws2.merge_cells("A1:F1")
ws2["A1"] = "REKAPITULASI HASIL PENGUJIAN BLACK BOX - SISTEM ATEKA"
ws2["A1"].font = Font(name=font_family, size=14, bold=True, color='0F1F3D')
ws2["A1"].alignment = Alignment(horizontal='left', vertical='center')
ws2.row_dimensions[1].height = 30

# Headers (Indonesian standard: No, Modul Pengujian, Jumlah Skenario, Jumlah Sesuai, Jumlah Tidak Sesuai, Persentase Kelayakan)
recap_headers = [
    "No", 
    "Modul Pengujian", 
    "Jumlah Skenario", 
    "Jumlah Sesuai", 
    "Jumlah Tidak Sesuai", 
    "Persentase Kelayakan"
]

recap_header_row = 3
for col_idx, header in enumerate(recap_headers, 1):
    cell = ws2.cell(row=recap_header_row, column=col_idx, value=header)
    cell.font = Font(name=font_family, size=11, bold=True, color='FFFFFF')
    cell.fill = navy_header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border
ws2.row_dimensions[recap_header_row].height = 28

# Add recap rows
recap_rows_data = [
    (1, "Autentikasi & Otorisasi", 4, 4, 0),
    (2, "Sinkronisasi Data (SIJUWAN Sync)", 3, 3, 0),
    (3, "Manajemen Akademik (Jurusan, Kelas, Angkatan, Mapel)", 5, 5, 0),
    (4, "Manajemen Soal (Bank Soal & Koleksi)", 4, 4, 0),
    (5, "Penjadwalan Ujian (Jadwal Wizard)", 4, 4, 0),
    (6, "Pelaksanaan Ujian Siswa (Mobile App)", 6, 6, 0),
    (7, "Analisis & Rekap Hasil Ujian", 3, 3, 0),
]

start_recap_row = 4
for idx, (no, name, sc_cnt, pass_cnt, fail_cnt) in enumerate(recap_rows_data):
    row_idx = start_recap_row + idx
    
    # Values (with dynamic Excel formula for Percentage)
    percentage_formula = f"=D{row_idx}/C{row_idx}"
    row_values = [no, name, sc_cnt, pass_cnt, fail_cnt, percentage_formula]
    
    for col_idx, val in enumerate(row_values, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        cell.font = Font(name=font_family, size=10)
        cell.border = thin_border
        
        # Alignment
        if col_idx in [1, 3, 4, 5, 6]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
            
        # Format percentage column
        if col_idx == 6:
            cell.number_format = '0.00%'
            cell.font = Font(name=font_family, size=10, bold=True, color='155724')
            cell.fill = pass_fill
            
    ws2.row_dimensions[row_idx].height = 24

# Add Total Row
total_row_idx = start_recap_row + len(recap_rows_data)
ws2.cell(row=total_row_idx, column=1, value="")
total_label = ws2.cell(row=total_row_idx, column=2, value="Total")
total_scenarios = ws2.cell(row=total_row_idx, column=3, value=f"=SUM(C4:C{total_row_idx-1})")
total_pass = ws2.cell(row=total_row_idx, column=4, value=f"=SUM(D4:D{total_row_idx-1})")
total_fail = ws2.cell(row=total_row_idx, column=5, value=f"=SUM(E4:E{total_row_idx-1})")
total_percentage = ws2.cell(row=total_row_idx, column=6, value=f"=D{total_row_idx}/C{total_row_idx}")

# Double bottom accounting border
total_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='double', color='000000')
)

for col_idx in range(1, 7):
    cell = ws2.cell(row=total_row_idx, column=col_idx)
    cell.font = Font(name=font_family, size=11, bold=True)
    cell.border = total_border
    cell.fill = module_header_fill
    if col_idx in [3, 4, 5, 6]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    else:
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
    if col_idx == 6:
        cell.number_format = '0.00%'
        cell.font = Font(name=font_family, size=11, bold=True, color='155724')

ws2.row_dimensions[total_row_idx].height = 26

# Column widths for Sheet 2
column_widths_s2 = {
    'A': 6,   # No
    'B': 45,  # Modul Pengujian
    'C': 18,  # Jumlah Skenario
    'D': 15,  # Jumlah Sesuai
    'E': 15,  # Jumlah Tidak Sesuai
    'F': 22   # Persentase Kelayakan
}
for col_letter, width in column_widths_s2.items():
    ws2.column_dimensions[col_letter].width = width

# Save the workbook
output_path = "d:\\FILE TUGAS AKHIR\\~APLIKASI\\Rekap_Blackbox_Testing_ATEKA.xlsx"
try:
    wb.save(output_path)
    print(f"Successfully generated Excel with codes removed at: {output_path}")
except Exception as e:
    print(f"Error saving workbook: {e}")
